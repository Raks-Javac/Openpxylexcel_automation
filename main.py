from pathlib import Path as ph
import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
def path_checker(file_path):
    file_processor = ph(file_path).exists()
    if(file_processor is True):
        print(f'{file_path} exists \n Processing excel file.....') 
    else:
        print(f'Oops file does not exist .....')
def excel_processor():
    for excel_iterator in range(2, sheet.max_row+1):
        cell = sheet.cell(excel_iterator,1)
        percentage_cell_prices = f'    ${float(cell.value/100)}'
        percentage_column = sheet.cell(excel_iterator , 4)
        percentage_column.value = percentage_cell_prices
        print(percentage_column.value)
        wb.save('transactions2.xlsx')

chart_values = Reference(
    min_row=2,
    max_row=sheet.max_row,
    max_col=4,
    min_col=4,
)
chart_diagrem = BarChart()
chart_diagrem.add_data(chart_values)
sheet.add_chart(chart_diagrem,'a2')
path_checker("transactions.xlsx")
excel_processor()


